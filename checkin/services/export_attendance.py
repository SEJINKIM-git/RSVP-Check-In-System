import csv
import io
import re

from django.http import HttpResponse
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from checkin.models import GuestParticipant, RSVPImportConfiguration, RegisteredParticipant
from checkin.services.import_rsvp import build_participant_answers, get_import_configuration_snapshot


ATTENDANCE_EXPORT_HEADERS = [
    "Type",
    "Name",
    "UNID",
    "Major",
    "Checked In",
    "Check-in Time",
]

RSVP_EXPORT_HEADERS = [
    "Submission Order",
    "Name",
    "UNID",
    "Major",
    "Checked In",
    "Check-in Time",
]

GROUPING_COLUMN_CANDIDATES = (
    "major",
    "department",
    "school",
    "program",
    "college",
    "team",
    "table",
    "group",
)
ATTENDANCE_GROUP_COLUMN_CANDIDATES = (
    "major",
    "department",
    "school",
    "program",
    "college",
)

XLSX_EXPORT_FILENAME = "rsvp_attendance_export.xlsx"
XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
HEADER_FONT = Font(bold=True)
UNSPECIFIED_GROUP_LABEL = "Undeclared"
OTHER_GROUP_LABEL = "Other"


def _format_datetime(value):
    if not value:
        return ""
    return timezone.localtime(value).strftime("%Y-%m-%d %H:%M:%S")


def _format_percentage(numerator, denominator):
    if not denominator:
        return "0.0%"
    return f"{(numerator / denominator) * 100:.1f}%"


def _format_column_list(columns):
    return ", ".join(columns) if columns else "None"


def _normalize_label(value):
    normalized = str(value or "").strip().lower()
    normalized = re.sub(r"[/_-]+", " ", normalized)
    normalized = re.sub(r"[^a-z0-9\s]", " ", normalized)
    return " ".join(normalized.split())


def _dedupe_preserve_order(values):
    seen = set()
    result = []
    for value in values:
        cleaned = str(value or "").strip()
        if not cleaned or cleaned in seen:
            continue
        seen.add(cleaned)
        result.append(cleaned)
    return result


def _collect_answer_keys(participants, preferred_columns=None):
    columns = []
    seen = set()

    for column in preferred_columns or []:
        cleaned = str(column or "").strip()
        if cleaned and cleaned not in seen:
            seen.add(cleaned)
            columns.append(cleaned)

    for participant in participants:
        for key in (participant.answers or {}).keys():
            cleaned = str(key or "").strip()
            if cleaned and cleaned not in seen:
                seen.add(cleaned)
                columns.append(cleaned)

    return columns


def _resolve_selected_columns(participants, configuration_snapshot, has_saved_configuration):
    configured_columns = _dedupe_preserve_order(configuration_snapshot.get("display_columns") or [])
    if has_saved_configuration and configured_columns:
        return configured_columns

    answer_keys = _collect_answer_keys(participants)
    if answer_keys:
        return answer_keys

    configured_columns = _dedupe_preserve_order(configuration_snapshot.get("imported_columns") or [])
    if configured_columns:
        return configured_columns

    return ["Name", "UNID", "Major"]


def _needs_unique_identifier_column(selected_columns, configuration_snapshot):
    if configuration_snapshot.get("unique_identifier_strategy") != "column":
        return True

    identifier_source = configuration_snapshot.get("unique_identifier_source", "")
    if not identifier_source:
        return True

    return identifier_source not in selected_columns


def _participant_answers(participant, configuration_snapshot):
    return build_participant_answers(participant, configuration_snapshot)


def _participant_value(participant, column_name, configuration_snapshot):
    return _participant_answers(participant, configuration_snapshot).get(column_name, "")


def _guess_guest_value_for_column(column_name, guest):
    normalized = _normalize_label(column_name)
    if "name" in normalized:
        return guest.name
    if any(token in normalized for token in ("unid", "student id", "uid", "identifier", "id number")):
        return guest.unid
    if any(token in normalized for token in ("major", "department", "school", "program", "college")):
        return guest.major
    return ""


def _normalize_group_value(value):
    cleaned = str(value or "").strip()
    if not cleaned or cleaned.lower() in {"none", "n/a", "-", "null", "na"}:
        return UNSPECIFIED_GROUP_LABEL
    if cleaned.lower() == "other":
        return OTHER_GROUP_LABEL
    return cleaned


def _group_sort_key(item):
    name, count = item
    if name == UNSPECIFIED_GROUP_LABEL:
        return (2, -count, name.lower())
    if name == OTHER_GROUP_LABEL:
        return (1, -count, name.lower())
    return (0, -count, name.lower())


def resolve_grouping_column(selected_columns, configuration_snapshot=None):
    normalized_columns = [
        (column_name, _normalize_label(column_name))
        for column_name in selected_columns or []
    ]
    configured_major_column = str(
        (configuration_snapshot or {}).get("major_column") or ""
    ).strip()

    if configured_major_column:
        normalized_major_column = _normalize_label(configured_major_column)
        for column_name, normalized_name in normalized_columns:
            if normalized_name == normalized_major_column:
                return column_name

    for candidate in GROUPING_COLUMN_CANDIDATES:
        for column_name, normalized_name in normalized_columns:
            if candidate in normalized_name:
                return column_name

    return ""


def _build_registered_headers(selected_columns, include_unique_identifier):
    headers = []
    if include_unique_identifier:
        headers.append("Unique Identifier")
    headers.extend(selected_columns)
    headers.extend(["Check-In Time", "Imported At", "Check-In Status"])
    return headers


def _build_registered_row(participant, selected_columns, include_unique_identifier, configuration_snapshot):
    row = []
    if include_unique_identifier:
        row.append(participant.unid)
    row.extend(
        _participant_value(participant, column_name, configuration_snapshot)
        for column_name in selected_columns
    )
    row.extend(
        [
            _format_datetime(participant.checkin_time),
            _format_datetime(participant.created_at),
            "Checked In" if participant.checked_in else "No Show",
        ]
    )
    return row


def _build_guest_headers():
    return ["Name", "UNID", "Major", "Created At", "Check-In Time", "Guest Status"]


def _build_guest_row(guest):
    return [
        guest.name,
        guest.unid,
        guest.major,
        _format_datetime(guest.created_at),
        _format_datetime(guest.checkin_time),
        "Checked In" if guest.checked_in else "Pending",
    ]


def _build_final_attendance_headers(selected_columns, include_unique_identifier):
    headers = ["Type"]
    if include_unique_identifier:
        headers.append("Unique Identifier")
    headers.extend(selected_columns)
    headers.extend(["Check-In Time", "Attendance Status"])
    return headers


def _build_final_registered_row(participant, selected_columns, include_unique_identifier, configuration_snapshot):
    row = ["Registered"]
    if include_unique_identifier:
        row.append(participant.unid)
    row.extend(
        _participant_value(participant, column_name, configuration_snapshot)
        for column_name in selected_columns
    )
    row.extend([_format_datetime(participant.checkin_time), "Present"])
    return row


def _build_final_guest_row(guest, selected_columns, include_unique_identifier):
    row = ["Guest"]
    if include_unique_identifier:
        row.append(guest.unid)
    row.extend(_guess_guest_value_for_column(column_name, guest) for column_name in selected_columns)
    row.extend(
        [
            _format_datetime(guest.checkin_time or guest.created_at),
            "Present" if guest.checked_in else "Pending",
        ]
    )
    return row


def _grouped_analysis_rows(participants, selected_columns, configuration_snapshot):
    grouping_column = resolve_grouping_column(selected_columns, configuration_snapshot)
    if not grouping_column:
        return ""

    grouped_stats = {}
    for participant in participants:
        group_value = _participant_value(participant, grouping_column, configuration_snapshot) or "Unspecified"
        stats = grouped_stats.setdefault(group_value, {"total": 0, "checked_in": 0})
        stats["total"] += 1
        if participant.checked_in:
            stats["checked_in"] += 1

    grouped_rows = []
    for group_value in sorted(grouped_stats):
        stats = grouped_stats[group_value]
        no_show_count = stats["total"] - stats["checked_in"]
        grouped_rows.append([group_value, stats["total"], stats["checked_in"], no_show_count])

    return grouping_column, grouped_rows


def build_current_attendance_group_rows(
    registered_participants,
    guest_participants,
    selected_columns,
    configuration_snapshot,
):
    normalized_columns = [
        (column_name, _normalize_label(column_name))
        for column_name in selected_columns or []
    ]
    configured_major_column = str(
        (configuration_snapshot or {}).get("major_column") or ""
    ).strip()
    grouping_column = ""

    if configured_major_column:
        normalized_major_column = _normalize_label(configured_major_column)
        for column_name, normalized_name in normalized_columns:
            if normalized_name == normalized_major_column:
                grouping_column = column_name
                break

    if not grouping_column:
        for candidate in ATTENDANCE_GROUP_COLUMN_CANDIDATES:
            for column_name, normalized_name in normalized_columns:
                if candidate in normalized_name:
                    grouping_column = column_name
                    break
            if grouping_column:
                break

    grouping_column = (
        grouping_column
        or configured_major_column
        or "Major"
    )
    grouped_counts = {}

    for participant in registered_participants:
        if not participant.checked_in:
            continue
        group_value = _participant_value(participant, grouping_column, configuration_snapshot)
        normalized_value = _normalize_group_value(group_value)
        grouped_counts[normalized_value] = grouped_counts.get(normalized_value, 0) + 1

    for guest in guest_participants:
        if not guest.checked_in:
            continue
        group_value = _guess_guest_value_for_column(grouping_column, guest)
        normalized_value = _normalize_group_value(group_value)
        grouped_counts[normalized_value] = grouped_counts.get(normalized_value, 0) + 1

    grouped_rows = sorted(grouped_counts.items(), key=_group_sort_key)
    return grouping_column, grouped_rows


def _style_header_row(worksheet):
    for cell in worksheet[1]:
        cell.font = HEADER_FONT


def _autosize_columns(worksheet):
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value_length = len(str(cell.value or ""))
            if value_length > max_length:
                max_length = value_length
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 48)


def _finalize_data_sheet(worksheet):
    _style_header_row(worksheet)
    worksheet.freeze_panes = "A2"
    _autosize_columns(worksheet)


def _add_summary_sheet(workbook, participants, guest_participants, selected_columns):
    worksheet = workbook.create_sheet("Summary")
    total_rsvp = len(participants)
    checked_in_rsvp = sum(1 for participant in participants if participant.checked_in)
    no_show_rsvp = total_rsvp - checked_in_rsvp
    guest_count = len(guest_participants)
    total_attendance = checked_in_rsvp + guest_count

    worksheet.append(["Metric", "Value"])
    worksheet.append(["Export Date", _format_datetime(timezone.now())])
    worksheet.append(["Total RSVP", total_rsvp])
    worksheet.append(["Checked-in RSVP", checked_in_rsvp])
    worksheet.append(["No-show RSVP", no_show_rsvp])
    worksheet.append(["Guest Count", guest_count])
    worksheet.append(["Total Attendance", total_attendance])
    worksheet.append(["RSVP Attendance Rate", _format_percentage(checked_in_rsvp, total_rsvp)])
    worksheet.append(["Selected imported column names", _format_column_list(selected_columns)])

    _finalize_data_sheet(worksheet)


def _add_registered_sheet(workbook, sheet_name, participants, selected_columns, include_unique_identifier, configuration_snapshot):
    worksheet = workbook.create_sheet(sheet_name)
    worksheet.append(_build_registered_headers(selected_columns, include_unique_identifier))

    for participant in participants:
        worksheet.append(
            _build_registered_row(
                participant,
                selected_columns,
                include_unique_identifier,
                configuration_snapshot,
            )
        )

    _finalize_data_sheet(worksheet)


def _add_guest_sheet(workbook, guest_participants):
    worksheet = workbook.create_sheet("Guest Participants")
    worksheet.append(_build_guest_headers())

    for guest in guest_participants:
        worksheet.append(_build_guest_row(guest))

    _finalize_data_sheet(worksheet)


def _add_final_attendance_sheet(workbook, checked_in_participants, guest_participants, selected_columns, include_unique_identifier, configuration_snapshot):
    worksheet = workbook.create_sheet("Final Attendance")
    worksheet.append(_build_final_attendance_headers(selected_columns, include_unique_identifier))

    for participant in checked_in_participants:
        worksheet.append(
            _build_final_registered_row(
                participant,
                selected_columns,
                include_unique_identifier,
                configuration_snapshot,
            )
        )

    for guest in guest_participants:
        worksheet.append(
            _build_final_guest_row(
                guest,
                selected_columns,
                include_unique_identifier,
            )
        )

    _finalize_data_sheet(worksheet)


def _add_analysis_sheet(workbook, participants, guest_participants, selected_columns, configuration_snapshot):
    worksheet = workbook.create_sheet("Data Analysis")
    total_rsvp = len(participants)
    checked_in_rsvp = sum(1 for participant in participants if participant.checked_in)
    no_show_rsvp = total_rsvp - checked_in_rsvp
    guest_count = len(guest_participants)
    total_attendance = checked_in_rsvp + guest_count

    worksheet.append(["Metric", "Value"])
    worksheet.append(["Total RSVP", total_rsvp])
    worksheet.append(["Checked-in RSVP", checked_in_rsvp])
    worksheet.append(["No-show RSVP", no_show_rsvp])
    worksheet.append(["Guest Count", guest_count])
    worksheet.append(["Total Attendance", total_attendance])
    worksheet.append(["RSVP Attendance Rate", _format_percentage(checked_in_rsvp, total_rsvp)])
    worksheet.append([])
    worksheet.append(["Attendance Type", "Count"])
    worksheet.append(["Registered", checked_in_rsvp])
    worksheet.append(["Guest", guest_count])
    worksheet.append([])
    worksheet.append(["RSVP Status", "Count"])
    worksheet.append(["Checked In", checked_in_rsvp])
    worksheet.append(["No Show", no_show_rsvp])

    grouped_analysis = _grouped_analysis_rows(participants, selected_columns, configuration_snapshot)
    worksheet.append([])
    if grouped_analysis:
        grouping_column, grouped_rows = grouped_analysis
        worksheet.append([grouping_column, "RSVP Count", "Checked-in Count", "No-show Count"])
        for grouped_row in grouped_rows:
            worksheet.append(grouped_row)
    else:
        worksheet.append(["Grouped Analysis"])
        worksheet.append(["No suitable selected import column found for grouped analysis."])

    _finalize_data_sheet(worksheet)


def build_attendance_csv_response():
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="final_attendance.csv"'

    writer = csv.writer(response)
    writer.writerow(ATTENDANCE_EXPORT_HEADERS)

    checked_in_registered = RegisteredParticipant.objects.filter(
        checked_in=True
    ).order_by("submission_order", "id")
    guest_participants = GuestParticipant.objects.all().order_by("-checkin_time", "-created_at", "id")

    for participant in checked_in_registered:
        writer.writerow(
            [
                "Registered",
                participant.name,
                participant.unid,
                participant.major,
                "Yes",
                _format_datetime(participant.checkin_time),
            ]
        )

    for guest in guest_participants:
        writer.writerow(
            [
                "Guest",
                guest.name,
                guest.unid,
                guest.major,
                "Yes" if guest.checked_in else "No",
                _format_datetime(guest.checkin_time),
            ]
        )

    return response


def build_rsvp_csv_response():
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="rsvp_participants.csv"'

    writer = csv.writer(response)
    writer.writerow(RSVP_EXPORT_HEADERS)

    participants = RegisteredParticipant.objects.all().order_by("submission_order", "id")
    for participant in participants:
        writer.writerow(
            [
                participant.submission_order,
                participant.name,
                participant.unid,
                participant.major,
                "Yes" if participant.checked_in else "No",
                _format_datetime(participant.checkin_time),
            ]
        )

    return response


def build_rsvp_xlsx_response():
    participants = list(RegisteredParticipant.objects.all().order_by("submission_order", "id"))
    guest_participants = list(GuestParticipant.objects.all().order_by("-checkin_time", "-created_at", "id"))
    checked_in_participants = [participant for participant in participants if participant.checked_in]
    no_show_participants = [participant for participant in participants if not participant.checked_in]

    has_saved_configuration = RSVPImportConfiguration.objects.exists()
    configuration_snapshot = get_import_configuration_snapshot()
    selected_columns = _resolve_selected_columns(
        participants,
        configuration_snapshot,
        has_saved_configuration,
    )
    include_unique_identifier = _needs_unique_identifier_column(selected_columns, configuration_snapshot)

    workbook = Workbook()
    workbook.remove(workbook.active)

    _add_summary_sheet(workbook, participants, guest_participants, selected_columns)
    _add_registered_sheet(
        workbook,
        "Registered Participants",
        participants,
        selected_columns,
        include_unique_identifier,
        configuration_snapshot,
    )
    _add_registered_sheet(
        workbook,
        "Checked-In Only",
        checked_in_participants,
        selected_columns,
        include_unique_identifier,
        configuration_snapshot,
    )
    _add_registered_sheet(
        workbook,
        "No-Show",
        no_show_participants,
        selected_columns,
        include_unique_identifier,
        configuration_snapshot,
    )
    _add_guest_sheet(workbook, guest_participants)
    _add_final_attendance_sheet(
        workbook,
        checked_in_participants,
        guest_participants,
        selected_columns,
        include_unique_identifier,
        configuration_snapshot,
    )
    _add_analysis_sheet(
        workbook,
        participants,
        guest_participants,
        selected_columns,
        configuration_snapshot,
    )

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(output.getvalue(), content_type=XLSX_CONTENT_TYPE)
    response["Content-Disposition"] = f'attachment; filename="{XLSX_EXPORT_FILENAME}"'
    return response
