import csv
import io
import logging
import re
from difflib import SequenceMatcher

from django.db import transaction
from django.db.models import Max

from checkin.models import RSVPImportConfiguration, RegisteredParticipant


logger = logging.getLogger(__name__)

IMPORT_SESSION_KEY = "pending_rsvp_import"
IMPORT_RESULT_SESSION_KEY = "rsvp_import_result"

GENERATE_INTERNAL_ID = "__generate_internal__"
NAME_TIMESTAMP_ID = "__name_timestamp__"
INTERNAL_ROW_NUMBER = "__row_number"

HEADER_SCAN_LIMIT = 10
HIGH_CONFIDENCE_THRESHOLD = 0.9
LOW_CONFIDENCE_THRESHOLD = 0.72

LEGACY_DEFAULT_COLUMNS = ["Name", "UNID", "Major"]
LEGACY_SEARCHABLE_COLUMNS = ["Name", "UNID", "Major"]

SUGGESTED_FIELD_NAMES = ("name", "major", "email", "timestamp")
FIELD_LABELS = {
    "name": "Name",
    "major": "Major",
    "email": "Email",
    "timestamp": "Timestamp",
}

FIELD_ALIASES = {
    "name": {
        "name",
        "full name",
        "student name",
        "participant name",
        "attendee name",
        "first and last name",
        "your name",
        "what is your name",
        "what is your full name",
        "legal name",
        "preferred name",
        "participant full name",
        "name of attendee",
        "이름",
    },
    "major": {
        "major",
        "major program",
        "program",
        "degree program",
        "field of study",
        "area of study",
        "department",
        "academic program",
        "concentration",
        "what is your major",
        "major or department",
        "전공",
    },
    "email": {
        "email",
        "email address",
        "e mail",
        "e mail address",
        "u email",
        "university email",
    },
    "timestamp": {
        "timestamp",
        "submission time",
        "submitted at",
        "submission date",
        "date submitted",
        "time submitted",
        "response timestamp",
    },
    "unique_identifier": {
        "unid",
        "u nid",
        "student id",
        "student number",
        "student uid",
        "university id",
        "university id number",
        "uid",
        "u id",
        "utah id",
        "campus id",
        "what is your unid",
        "what is your student id",
    },
}


def _normalize_header(value):
    normalized = str(value or "").strip().lower()
    normalized = re.sub(r"[/_-]+", " ", normalized)
    normalized = re.sub(r"[^a-z0-9\u3131-\u318e\uac00-\ud7a3\s]", " ", normalized)
    return " ".join(normalized.split())


def _clean_cell(value):
    if value is None:
        return ""
    return str(value).strip()


def _dedupe_preserve_order(values):
    seen = set()
    result = []
    for value in values:
        cleaned = _clean_cell(value)
        if not cleaned or cleaned in seen:
            continue
        seen.add(cleaned)
        result.append(cleaned)
    return result


def _make_unique_headers(header_cells):
    unique_headers = []
    seen = {}

    for index, value in header_cells:
        header = _clean_cell(value) or f"Column {index + 1}"
        count = seen.get(header, 0) + 1
        seen[header] = count
        if count > 1:
            header = f"{header} ({count})"
        unique_headers.append((index, header))

    return unique_headers


def _decode_uploaded_file(uploaded_file):
    uploaded_file.seek(0)
    raw_content = uploaded_file.read()

    if isinstance(raw_content, str):
        return raw_content

    for encoding in ("utf-8-sig", "utf-8", "cp1252"):
        try:
            return raw_content.decode(encoding)
        except UnicodeDecodeError:
            continue

    raise ValueError("The uploaded file could not be decoded as text.")


def _parse_csv_rows(uploaded_file):
    file_text = _decode_uploaded_file(uploaded_file)
    csv_buffer = io.StringIO(file_text, newline="")

    try:
        reader = csv.reader(csv_buffer)
        return [list(row) for row in reader]
    except csv.Error as error:
        raise ValueError(f"Could not read CSV file: {error}") from error


def _parse_xlsx_rows(uploaded_file):
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise ValueError("XLSX uploads require the openpyxl package to be installed.") from error

    uploaded_file.seek(0)
    try:
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    except Exception as error:
        raise ValueError(f"Could not read XLSX file: {error}") from error

    worksheet = workbook.active
    return [list(row) for row in worksheet.iter_rows(values_only=True)]


def _detect_header_row(raw_rows):
    if not raw_rows:
        raise ValueError("The uploaded file is empty.")

    scan_rows = raw_rows[:HEADER_SCAN_LIMIT]
    for index, row in enumerate(scan_rows):
        non_empty_values = [_clean_cell(value) for value in row if _clean_cell(value)]
        if len(non_empty_values) >= 2:
            return index

    for index, row in enumerate(raw_rows):
        if any(_clean_cell(value) for value in row):
            return index

    raise ValueError("The uploaded file does not contain any readable rows.")


def _rows_to_records(raw_rows):
    header_row_index = _detect_header_row(raw_rows)
    raw_header_row = raw_rows[header_row_index]
    non_empty_header_cells = [
        (index, value)
        for index, value in enumerate(raw_header_row)
        if _clean_cell(value)
    ]
    if not non_empty_header_cells:
        raise ValueError("The uploaded file is missing a usable header row.")

    indexed_headers = _make_unique_headers(non_empty_header_cells)
    headers = [header for _, header in indexed_headers]
    rows = []

    for row_index, raw_row in enumerate(raw_rows[header_row_index + 1 :], start=header_row_index + 2):
        record = {
            header: _clean_cell(raw_row[column_index] if column_index < len(raw_row) else "")
            for column_index, header in indexed_headers
        }
        if any(record.values()):
            record[INTERNAL_ROW_NUMBER] = row_index
            rows.append(record)

    return headers, rows, header_row_index + 1


def parse_rsvp_file(uploaded_file):
    if not uploaded_file:
        raise ValueError("Please choose a CSV or XLSX file to import.")

    file_name = (uploaded_file.name or "").lower()
    if file_name.endswith(".csv"):
        raw_rows = _parse_csv_rows(uploaded_file)
    elif file_name.endswith(".xlsx"):
        raw_rows = _parse_xlsx_rows(uploaded_file)
    elif file_name.endswith(".xls"):
        raise ValueError("Legacy XLS files are not supported. Please save the file as CSV or XLSX.")
    else:
        raise ValueError("Only CSV and XLSX uploads are supported.")

    return _rows_to_records(raw_rows)


def _score_header_for_field(header, field_name):
    normalized_header = _normalize_header(header)
    if not normalized_header:
        return 0

    aliases = FIELD_ALIASES[field_name]
    if normalized_header in aliases:
        return 1

    header_tokens = set(normalized_header.split())
    best_score = 0
    for alias in aliases:
        alias_tokens = set(alias.split())
        if not alias_tokens:
            continue
        if alias in normalized_header:
            best_score = max(best_score, 0.92)
        elif alias_tokens.issubset(header_tokens):
            best_score = max(best_score, 0.88)
        elif header_tokens.issubset(alias_tokens):
            best_score = max(best_score, 0.78)
        else:
            best_score = max(best_score, SequenceMatcher(None, normalized_header, alias).ratio())

    return round(best_score, 2)


def detect_column_mapping(headers):
    matches = {}

    for field_name in (*SUGGESTED_FIELD_NAMES, "unique_identifier"):
        scored_headers = sorted(
            (
                (_score_header_for_field(header, field_name), header)
                for header in headers
            ),
            reverse=True,
        )
        if not scored_headers:
            continue

        score, header = scored_headers[0]
        if score >= LOW_CONFIDENCE_THRESHOLD:
            matches[field_name] = {
                "header": header,
                "confidence": score,
                "label": "Unique Identifier" if field_name == "unique_identifier" else FIELD_LABELS[field_name],
            }

    return {"matches": matches}


def _is_identifier_like_header(header):
    normalized_header = _normalize_header(header)
    header_tokens = set(normalized_header.split())
    if {"id", "uid", "unid", "number"} & header_tokens:
        return True
    return any(
        phrase in normalized_header
        for phrase in (
            "student id",
            "student number",
            "university id",
            "campus id",
            "utah id",
        )
    )


def _guess_unique_identifier_selection(detection):
    matches = detection["matches"]
    unique_identifier_match = matches.get("unique_identifier")
    if unique_identifier_match and _is_identifier_like_header(unique_identifier_match["header"]):
        return unique_identifier_match["header"]

    email_match = matches.get("email")
    if email_match:
        return email_match["header"]

    if matches.get("name") and matches.get("timestamp"):
        return NAME_TIMESTAMP_ID

    return GENERATE_INTERNAL_ID


def _default_display_columns(headers, matches, unique_identifier_selection):
    recommended = [
        matches.get("name", {}).get("header", ""),
        unique_identifier_selection
        if unique_identifier_selection not in {GENERATE_INTERNAL_ID, NAME_TIMESTAMP_ID}
        else "",
        matches.get("major", {}).get("header", ""),
        matches.get("email", {}).get("header", ""),
    ]
    recommended.extend(headers[:4])
    return _dedupe_preserve_order(recommended)[:4] or headers[:4]


def build_default_review_settings(headers, detection):
    matches = detection["matches"]
    unique_identifier_selection = _guess_unique_identifier_selection(detection)

    display_columns = _default_display_columns(headers, matches, unique_identifier_selection)
    searchable_columns = _dedupe_preserve_order(
        [
            unique_identifier_selection
            if unique_identifier_selection not in {GENERATE_INTERNAL_ID, NAME_TIMESTAMP_ID}
            else "",
            matches.get("email", {}).get("header", ""),
            matches.get("name", {}).get("header", ""),
            matches.get("major", {}).get("header", ""),
            *display_columns,
        ]
    ) or display_columns

    return {
        "unique_identifier_selection": unique_identifier_selection,
        "unique_identifier_strategy": RSVPImportConfiguration.UNIQUE_IDENTIFIER_COLUMN
        if unique_identifier_selection not in {GENERATE_INTERNAL_ID, NAME_TIMESTAMP_ID}
        else RSVPImportConfiguration.UNIQUE_IDENTIFIER_INTERNAL,
        "unique_identifier_source": (
            unique_identifier_selection
            if unique_identifier_selection not in {GENERATE_INTERNAL_ID, NAME_TIMESTAMP_ID}
            else ""
        ),
        "name_column": matches.get("name", {}).get("header", ""),
        "major_column": matches.get("major", {}).get("header", ""),
        "email_column": matches.get("email", {}).get("header", ""),
        "timestamp_column": matches.get("timestamp", {}).get("header", ""),
        "display_columns": display_columns,
        "searchable_columns": searchable_columns,
    }


def _valid_selected_headers(values, headers):
    header_set = set(headers)
    return _dedupe_preserve_order(value for value in values if value in header_set)


def normalize_review_settings(review_settings, headers, detection):
    defaults = build_default_review_settings(headers, detection)
    review_settings = review_settings or {}
    header_set = set(headers)

    name_column = review_settings.get("name_column", defaults["name_column"])
    major_column = review_settings.get("major_column", defaults["major_column"])
    email_column = review_settings.get("email_column", defaults["email_column"])
    timestamp_column = review_settings.get("timestamp_column", defaults["timestamp_column"])

    name_column = name_column if name_column in header_set else defaults["name_column"]
    major_column = major_column if major_column in header_set else defaults["major_column"]
    email_column = email_column if email_column in header_set else defaults["email_column"]
    timestamp_column = (
        timestamp_column if timestamp_column in header_set else defaults["timestamp_column"]
    )

    unique_identifier_selection = review_settings.get(
        "unique_identifier_selection",
        defaults["unique_identifier_selection"],
    )
    if unique_identifier_selection == NAME_TIMESTAMP_ID and not (name_column and timestamp_column):
        unique_identifier_selection = defaults["unique_identifier_selection"]
    if unique_identifier_selection not in header_set and unique_identifier_selection not in {
        GENERATE_INTERNAL_ID,
        NAME_TIMESTAMP_ID,
    }:
        unique_identifier_selection = defaults["unique_identifier_selection"]

    if unique_identifier_selection == NAME_TIMESTAMP_ID:
        unique_identifier_strategy = RSVPImportConfiguration.UNIQUE_IDENTIFIER_NAME_TIMESTAMP
        unique_identifier_source = ""
    elif unique_identifier_selection == GENERATE_INTERNAL_ID:
        unique_identifier_strategy = RSVPImportConfiguration.UNIQUE_IDENTIFIER_INTERNAL
        unique_identifier_source = ""
    else:
        unique_identifier_strategy = RSVPImportConfiguration.UNIQUE_IDENTIFIER_COLUMN
        unique_identifier_source = unique_identifier_selection

    display_columns = _valid_selected_headers(
        review_settings.get("display_columns", defaults["display_columns"]),
        headers,
    )
    if not display_columns:
        display_columns = defaults["display_columns"] or headers[:4]

    searchable_columns = _valid_selected_headers(
        review_settings.get("searchable_columns", defaults["searchable_columns"]),
        headers,
    )
    if not searchable_columns:
        searchable_columns = display_columns[:]

    return {
        "unique_identifier_selection": unique_identifier_selection,
        "unique_identifier_strategy": unique_identifier_strategy,
        "unique_identifier_source": unique_identifier_source,
        "name_column": name_column,
        "major_column": major_column,
        "email_column": email_column,
        "timestamp_column": timestamp_column,
        "display_columns": display_columns,
        "searchable_columns": searchable_columns,
    }


def get_unique_identifier_label(configuration):
    strategy = configuration.get("unique_identifier_strategy")
    if strategy == RSVPImportConfiguration.UNIQUE_IDENTIFIER_NAME_TIMESTAMP:
        return "Name + Timestamp"
    if strategy == RSVPImportConfiguration.UNIQUE_IDENTIFIER_INTERNAL:
        return "Generated RSVP ID"
    return configuration.get("unique_identifier_source") or "Unique Identifier"


def build_unique_identifier_options(headers, review_settings):
    options = [{"value": header, "label": header} for header in headers]
    if review_settings.get("name_column") and review_settings.get("timestamp_column"):
        options.append({"value": NAME_TIMESTAMP_ID, "label": "Name + Timestamp"})
    options.append(
        {
            "value": GENERATE_INTERNAL_ID,
            "label": "Generate internal RSVP ID automatically",
        }
    )
    return options


def _build_preview_identifier(row, review_settings):
    strategy = review_settings["unique_identifier_strategy"]
    row_number = row[INTERNAL_ROW_NUMBER]

    if strategy == RSVPImportConfiguration.UNIQUE_IDENTIFIER_INTERNAL:
        return {
            "display_value": "Generated automatically",
            "normalized_value": f"generated-{row_number}",
            "errors": [],
        }

    if strategy == RSVPImportConfiguration.UNIQUE_IDENTIFIER_NAME_TIMESTAMP:
        name_value = _clean_cell(row.get(review_settings["name_column"]))
        timestamp_value = _clean_cell(row.get(review_settings["timestamp_column"]))
        if not name_value or not timestamp_value:
            return {
                "display_value": "",
                "normalized_value": "",
                "errors": ["Name + Timestamp requires both mapped values in each row."],
            }
        identifier_value = f"{name_value} | {timestamp_value}"
        return {
            "display_value": identifier_value,
            "normalized_value": identifier_value.lower(),
            "errors": [],
        }

    identifier_value = _clean_cell(row.get(review_settings["unique_identifier_source"]))
    if not identifier_value:
        return {
            "display_value": "",
            "normalized_value": "",
            "errors": [
                f"Unique identifier column '{review_settings['unique_identifier_source']}' is blank."
            ],
        }

    return {
        "display_value": identifier_value,
        "normalized_value": identifier_value.lower(),
        "errors": [],
    }


def build_import_preview(rows, headers, review_settings):
    existing_identifiers = {
        (identifier or "").strip().lower()
        for identifier in RegisteredParticipant.objects.values_list("unid", flat=True)
    }
    seen_identifiers_in_file = set()
    preview_rows = []
    valid_count = 0
    invalid_count = 0
    duplicate_count = 0
    missing_identifier_count = 0

    for row in rows:
        identifier_info = _build_preview_identifier(row, review_settings)
        row_errors = list(identifier_info["errors"])
        normalized_identifier = identifier_info["normalized_value"]

        if identifier_info["errors"]:
            missing_identifier_count += 1

        if normalized_identifier:
            if normalized_identifier in existing_identifiers:
                row_errors.append("This unique identifier already exists in the RSVP list.")
                duplicate_count += 1
            elif normalized_identifier in seen_identifiers_in_file:
                row_errors.append("This unique identifier appears more than once in the uploaded file.")
                duplicate_count += 1

        if row_errors:
            invalid_count += 1
        else:
            valid_count += 1
            seen_identifiers_in_file.add(normalized_identifier)

        preview_rows.append(
            {
                "row_number": row[INTERNAL_ROW_NUMBER],
                "cells": [{"header": header, "value": row.get(header, "")} for header in headers],
                "is_valid": not row_errors,
                "errors": row_errors,
            }
        )

    return {
        "total_rows": len(rows),
        "valid_count": valid_count,
        "invalid_count": invalid_count,
        "duplicate_count": duplicate_count,
        "missing_identifier_count": missing_identifier_count,
        "preview_rows": preview_rows[:10],
    }


def build_import_review(headers, rows, detection, review_settings):
    normalized_settings = normalize_review_settings(review_settings, headers, detection)
    return {
        "detected_columns": headers,
        "header_count": len(headers),
        "review_settings": normalized_settings,
        "identifier_label": get_unique_identifier_label(normalized_settings),
        "identifier_options": build_unique_identifier_options(headers, normalized_settings),
        "preview": build_import_preview(rows, headers, normalized_settings),
    }


def prepare_rsvp_import(uploaded_file, review_settings=None):
    result = {
        "headers": [],
        "rows": [],
        "header_row_number": 0,
        "detection": {"matches": {}},
        "review": None,
        "errors": [],
    }

    try:
        headers, rows, header_row_number = parse_rsvp_file(uploaded_file)
    except ValueError as error:
        result["errors"].append(str(error))
        return result

    detection = detect_column_mapping(headers)
    review = build_import_review(headers, rows, detection, review_settings)

    result.update(
        {
            "headers": headers,
            "rows": rows,
            "header_row_number": header_row_number,
            "detection": detection,
            "review": review,
        }
    )
    return result


def save_import_configuration(review_settings, headers):
    configuration, _ = RSVPImportConfiguration.objects.get_or_create(pk=1)
    configuration.imported_columns = headers[:]
    configuration.display_columns = review_settings["display_columns"]
    configuration.searchable_columns = review_settings["searchable_columns"]
    configuration.unique_identifier_strategy = review_settings["unique_identifier_strategy"]
    configuration.unique_identifier_source = review_settings["unique_identifier_source"]
    configuration.name_column = review_settings["name_column"]
    configuration.major_column = review_settings["major_column"]
    configuration.email_column = review_settings["email_column"]
    configuration.timestamp_column = review_settings["timestamp_column"]
    configuration.save()
    return configuration


def get_import_configuration_snapshot():
    configuration = RSVPImportConfiguration.objects.order_by("pk").first()
    if not configuration:
        return {
            "imported_columns": LEGACY_DEFAULT_COLUMNS[:],
            "display_columns": LEGACY_DEFAULT_COLUMNS[:],
            "searchable_columns": LEGACY_SEARCHABLE_COLUMNS[:],
            "unique_identifier_strategy": RSVPImportConfiguration.UNIQUE_IDENTIFIER_COLUMN,
            "unique_identifier_source": "UNID",
            "name_column": "Name",
            "major_column": "Major",
            "email_column": "Email",
            "timestamp_column": "",
            "identifier_label": "UNID",
        }

    snapshot = {
        "imported_columns": _dedupe_preserve_order(configuration.imported_columns or []),
        "display_columns": _dedupe_preserve_order(configuration.display_columns or []),
        "searchable_columns": _dedupe_preserve_order(configuration.searchable_columns or []),
        "unique_identifier_strategy": configuration.unique_identifier_strategy,
        "unique_identifier_source": configuration.unique_identifier_source,
        "name_column": configuration.name_column,
        "major_column": configuration.major_column,
        "email_column": configuration.email_column,
        "timestamp_column": configuration.timestamp_column,
    }
    if not snapshot["imported_columns"]:
        snapshot["imported_columns"] = LEGACY_DEFAULT_COLUMNS[:]
    if not snapshot["display_columns"]:
        snapshot["display_columns"] = snapshot["imported_columns"][:3] or LEGACY_DEFAULT_COLUMNS[:]
    if not snapshot["searchable_columns"]:
        snapshot["searchable_columns"] = snapshot["display_columns"][:]
    snapshot["identifier_label"] = get_unique_identifier_label(snapshot)
    return snapshot


def build_participant_answers(participant, configuration=None):
    configuration = configuration or get_import_configuration_snapshot()
    answers = dict(participant.answers or {})

    if participant.name:
        answers.setdefault("Name", participant.name)
    if participant.unid:
        answers.setdefault("UNID", participant.unid)
    if participant.major:
        answers.setdefault("Major", participant.major)
    if participant.email:
        answers.setdefault("Email", participant.email)

    if configuration.get("name_column") and participant.name:
        answers.setdefault(configuration["name_column"], participant.name)
    if configuration.get("major_column") and participant.major:
        answers.setdefault(configuration["major_column"], participant.major)
    if configuration.get("email_column") and participant.email:
        answers.setdefault(configuration["email_column"], participant.email)
    if (
        configuration.get("unique_identifier_strategy") == RSVPImportConfiguration.UNIQUE_IDENTIFIER_COLUMN
        and configuration.get("unique_identifier_source")
        and participant.unid
    ):
        answers.setdefault(configuration["unique_identifier_source"], participant.unid)

    return answers


def _import_identifier_for_row(row, review_settings, next_submission_order):
    strategy = review_settings["unique_identifier_strategy"]
    row_number = row[INTERNAL_ROW_NUMBER]

    if strategy == RSVPImportConfiguration.UNIQUE_IDENTIFIER_INTERNAL:
        generated_identifier = f"rsvp-{next_submission_order:05d}"
        return {
            "display_value": generated_identifier,
            "normalized_value": generated_identifier.lower(),
            "errors": [],
        }

    if strategy == RSVPImportConfiguration.UNIQUE_IDENTIFIER_NAME_TIMESTAMP:
        name_value = _clean_cell(row.get(review_settings["name_column"]))
        timestamp_value = _clean_cell(row.get(review_settings["timestamp_column"]))
        if not name_value or not timestamp_value:
            return {
                "display_value": "",
                "normalized_value": "",
                "errors": [
                    f"Row {row_number} skipped: Name + Timestamp requires both mapped values."
                ],
            }
        identifier_value = f"{name_value} | {timestamp_value}"
        return {
            "display_value": identifier_value,
            "normalized_value": identifier_value.lower(),
            "errors": [],
        }

    identifier_value = _clean_cell(row.get(review_settings["unique_identifier_source"]))
    if not identifier_value:
        return {
            "display_value": "",
            "normalized_value": "",
            "errors": [
                f"Row {row_number} skipped: unique identifier column '{review_settings['unique_identifier_source']}' is blank."
            ],
        }

    return {
        "display_value": identifier_value,
        "normalized_value": identifier_value.lower(),
        "errors": [],
    }


def import_rsvp_rows(rows, headers, detection, review_settings):
    normalized_settings = normalize_review_settings(review_settings, headers, detection)
    summary = {
        "imported_count": 0,
        "skipped_count": 0,
        "duplicate_identifiers": [],
        "errors": [],
        "identifier_label": get_unique_identifier_label(normalized_settings),
        "total_rows": len(rows),
        "valid_count": 0,
        "invalid_count": 0,
        "missing_identifier_count": 0,
    }

    existing_identifiers = {
        (identifier or "").strip().lower()
        for identifier in RegisteredParticipant.objects.values_list("unid", flat=True)
    }
    seen_identifiers_in_file = set()
    next_submission_order = (
        RegisteredParticipant.objects.aggregate(max_order=Max("submission_order"))["max_order"] or 0
    )
    participants_to_create = []

    for row in rows:
        candidate_submission_order = next_submission_order + 1
        identifier_info = _import_identifier_for_row(
            row,
            normalized_settings,
            candidate_submission_order,
        )
        if identifier_info["errors"]:
            summary["skipped_count"] += 1
            summary["missing_identifier_count"] += 1
            summary["errors"].extend(identifier_info["errors"])
            continue

        normalized_identifier = identifier_info["normalized_value"]
        if normalized_identifier in existing_identifiers or normalized_identifier in seen_identifiers_in_file:
            summary["skipped_count"] += 1
            if identifier_info["display_value"] not in summary["duplicate_identifiers"]:
                summary["duplicate_identifiers"].append(identifier_info["display_value"])
            continue

        answers = {
            header: row.get(header, "")
            for header in headers
        }
        name_value = _clean_cell(row.get(normalized_settings["name_column"]))
        major_value = _clean_cell(row.get(normalized_settings["major_column"]))
        email_value = _clean_cell(row.get(normalized_settings["email_column"]))

        participant_display_name = (
            name_value
            or identifier_info["display_value"]
            or f"Imported Participant {candidate_submission_order}"
        )

        participants_to_create.append(
            RegisteredParticipant(
                submission_order=candidate_submission_order,
                name=participant_display_name,
                unid=normalized_identifier,
                major=major_value,
                email=email_value,
                answers=answers,
            )
        )
        seen_identifiers_in_file.add(normalized_identifier)
        next_submission_order = candidate_submission_order

    try:
        with transaction.atomic():
            RegisteredParticipant.objects.bulk_create(participants_to_create)
            save_import_configuration(normalized_settings, headers)
    except Exception as error:
        summary["errors"].append(f"Import failed while saving rows: {error}")
        summary["skipped_count"] = len(rows)
        summary["invalid_count"] = len(rows)
        return summary

    summary["imported_count"] = len(participants_to_create)
    summary["valid_count"] = len(participants_to_create)
    summary["invalid_count"] = summary["skipped_count"]
    return summary


def import_rsvp_file(uploaded_file, review_settings=None):
    summary = {
        "imported_count": 0,
        "skipped_count": 0,
        "duplicate_identifiers": [],
        "errors": [],
        "identifier_label": "Unique Identifier",
        "total_rows": 0,
        "valid_count": 0,
        "invalid_count": 0,
        "missing_identifier_count": 0,
    }

    prepared_import = prepare_rsvp_import(uploaded_file, review_settings=review_settings)
    if prepared_import["errors"]:
        summary["errors"].extend(prepared_import["errors"])
        return summary

    logger.debug("Detected RSVP import headers: %s", prepared_import["headers"])
    return import_rsvp_rows(
        prepared_import["rows"],
        prepared_import["headers"],
        prepared_import["detection"],
        prepared_import["review"]["review_settings"],
    )
