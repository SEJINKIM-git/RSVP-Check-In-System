import json
from io import BytesIO
from unittest.mock import patch

from django.core.files.uploadedfile import SimpleUploadedFile
from django.db.utils import OperationalError
from django.test import TestCase
from django.utils import timezone

from openpyxl import Workbook, load_workbook

from checkin.forms import DEFAULT_GUEST_MAJORS, GuestCheckInForm
from checkin.models import GuestParticipant, RSVPImportConfiguration, RegisteredParticipant
from checkin.services.import_rsvp import (
    IMPORT_SESSION_KEY,
    NAME_TIMESTAMP_ID,
    get_import_configuration_snapshot,
    import_rsvp_file,
    prepare_rsvp_import,
)


def build_xlsx_upload(name, rows):
    workbook = Workbook()
    worksheet = workbook.active
    for row in rows:
        worksheet.append(row)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return SimpleUploadedFile(
        name,
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


class RSVPImportServiceTests(TestCase):
    def test_prepare_import_detects_title_row_and_flexible_headers_in_csv(self):
        uploaded_file = SimpleUploadedFile(
            "rsvp.csv",
            "\n".join(
                [
                    "Table Seating Assignment - Start UP Sprint 2026",
                    "Timestamp,Full Name,Student ID,Dietary Restrictions,Department",
                    "2026-05-01 09:00,Alice Kim,U1234567,None,Computer Science",
                ]
            ).encode("utf-8"),
            content_type="text/csv",
        )

        prepared_import = prepare_rsvp_import(uploaded_file)

        self.assertEqual(prepared_import["errors"], [])
        self.assertEqual(prepared_import["header_row_number"], 2)
        self.assertEqual(
            prepared_import["headers"],
            [
                "Timestamp",
                "Full Name",
                "Student ID",
                "Dietary Restrictions",
                "Department",
            ],
        )
        self.assertEqual(
            prepared_import["review"]["review_settings"]["unique_identifier_selection"],
            "Student ID",
        )
        self.assertEqual(prepared_import["review"]["preview"]["valid_count"], 1)

    def test_prepare_import_suggests_name_timestamp_when_no_identifier_column_exists(self):
        uploaded_file = SimpleUploadedFile(
            "rsvp.csv",
            (
                "Timestamp,What is your full name?,Will you attend?\n"
                "2026-05-01 09:00,Alice Kim,Yes\n"
            ).encode("utf-8"),
            content_type="text/csv",
        )

        prepared_import = prepare_rsvp_import(uploaded_file)

        self.assertEqual(prepared_import["errors"], [])
        self.assertEqual(
            prepared_import["review"]["review_settings"]["unique_identifier_selection"],
            NAME_TIMESTAMP_ID,
        )

    def test_import_supports_xlsx_files_with_flexible_headers(self):
        uploaded_file = build_xlsx_upload(
            "rsvp.xlsx",
            [
                ["Start UP Sprint RSVP Responses"],
                ["Submitted At", "Participant Full Name", "Email Address", "Program"],
                ["2026-05-01 09:00", "Alice Kim", "alice@example.com", "Finance"],
                ["2026-05-01 09:02", "Brian Lee", "brian@example.com", "Marketing"],
            ],
        )

        summary = import_rsvp_file(uploaded_file)

        self.assertEqual(summary["imported_count"], 2)
        self.assertEqual(summary["errors"], [])
        self.assertEqual(
            list(
                RegisteredParticipant.objects.order_by("submission_order").values_list(
                    "name",
                    "unid",
                    "major",
                    "email",
                )
            ),
            [
                ("Alice Kim", "alice@example.com", "Finance", "alice@example.com"),
                ("Brian Lee", "brian@example.com", "Marketing", "brian@example.com"),
            ],
        )

    def test_import_can_generate_internal_ids_when_no_reliable_identifier_exists(self):
        uploaded_file = SimpleUploadedFile(
            "rsvp.csv",
            (
                "Favorite Color,Preferred Workshop\n"
                "Red,AI Product Design\n"
                "Blue,Startup Finance\n"
            ).encode("utf-8"),
            content_type="text/csv",
        )

        summary = import_rsvp_file(
            uploaded_file,
            review_settings={
                "unique_identifier_selection": "__generate_internal__",
                "display_columns": ["Favorite Color", "Preferred Workshop"],
                "searchable_columns": ["Favorite Color"],
            },
        )

        self.assertEqual(summary["imported_count"], 2)
        self.assertEqual(summary["errors"], [])
        participants = list(RegisteredParticipant.objects.order_by("submission_order"))
        self.assertTrue(all(participant.unid.startswith("rsvp-") for participant in participants))
        self.assertEqual(
            participants[0].answers["Preferred Workshop"],
            "AI Product Design",
        )

    def test_import_skips_duplicate_selected_identifiers(self):
        RegisteredParticipant.objects.create(
            submission_order=1,
            name="Existing Student",
            unid="u1234567",
            major="Economics",
        )

        uploaded_file = SimpleUploadedFile(
            "rsvp.csv",
            (
                "Full Name,Student ID,Department\n"
                "Alice Kim,U1234567,Finance\n"
                "Brian Lee,U1234567,Marketing\n"
            ).encode("utf-8"),
            content_type="text/csv",
        )

        summary = import_rsvp_file(uploaded_file)

        self.assertEqual(summary["imported_count"], 0)
        self.assertEqual(summary["skipped_count"], 2)
        self.assertEqual(summary["duplicate_identifiers"], ["U1234567"])

    def test_reports_legacy_xls_files_as_unsupported(self):
        uploaded_file = SimpleUploadedFile(
            "rsvp.xls",
            b"fake-xls-content",
            content_type="application/vnd.ms-excel",
        )

        summary = import_rsvp_file(uploaded_file)

        self.assertEqual(summary["imported_count"], 0)
        self.assertEqual(
            summary["errors"],
            ["Legacy XLS files are not supported. Please save the file as CSV or XLSX."],
        )


class ImportReviewFlowTests(TestCase):
    def test_upload_redirects_to_review_and_waits_for_confirmation(self):
        uploaded_file = SimpleUploadedFile(
            "rsvp.csv",
            (
                "Timestamp,Full Name,Student ID,Email,Major\n"
                "2026-05-01 09:00,Alice Kim,U1234567,alice@example.com,Finance\n"
            ).encode("utf-8"),
            content_type="text/csv",
        )

        response = self.client.post("/checkin/import/", {"rsvp_file": uploaded_file})

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], "/checkin/import/review/")
        self.assertEqual(RegisteredParticipant.objects.count(), 0)
        self.assertIn(IMPORT_SESSION_KEY, self.client.session)

    def test_review_confirm_import_persists_answers_and_configuration(self):
        uploaded_file = SimpleUploadedFile(
            "rsvp.csv",
            (
                "Timestamp,Full Name,Student ID,Email,Dietary Restrictions,Will you attend?\n"
                "2026-05-01 09:00,Alice Kim,U1234567,alice@example.com,None,Yes\n"
                "2026-05-01 09:05,Brian Lee,U2345678,brian@example.com,Vegetarian,Yes\n"
            ).encode("utf-8"),
            content_type="text/csv",
        )

        response = self.client.post("/checkin/import/", {"rsvp_file": uploaded_file})
        self.assertEqual(response.status_code, 302)

        response = self.client.post(
            "/checkin/import/review/",
            {
                "review_action": "confirm",
                "unique_identifier_selection": "Student ID",
                "name_column": "Full Name",
                "major_column": "",
                "email_column": "Email",
                "timestamp_column": "Timestamp",
                "display_columns": ["Full Name", "Email", "Dietary Restrictions"],
                "searchable_columns": ["Full Name", "Email", "Will you attend?"],
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(RegisteredParticipant.objects.count(), 2)

        participant = RegisteredParticipant.objects.get(unid="u1234567")
        self.assertEqual(participant.name, "Alice Kim")
        self.assertEqual(participant.email, "alice@example.com")
        self.assertEqual(participant.answers["Dietary Restrictions"], "None")
        self.assertEqual(
            participant.answers,
            {
                "Full Name": "Alice Kim",
                "Email": "alice@example.com",
                "Dietary Restrictions": "None",
            },
        )
        self.assertNotIn("Timestamp", participant.answers)
        self.assertNotIn("Student ID", participant.answers)
        self.assertNotIn("Will you attend?", participant.answers)

        configuration = RSVPImportConfiguration.objects.get(pk=1)
        self.assertEqual(
            configuration.imported_columns,
            ["Full Name", "Email", "Dietary Restrictions"],
        )
        self.assertEqual(configuration.unique_identifier_source, "Student ID")
        self.assertEqual(configuration.name_column, "Full Name")
        self.assertEqual(
            configuration.display_columns,
            ["Full Name", "Email", "Dietary Restrictions"],
        )
        self.assertEqual(configuration.searchable_columns, ["Full Name", "Email"])

        self.assertContains(response, "Imported Rows")
        self.assertContains(response, "2")

    def test_registered_page_uses_selected_display_and_search_columns(self):
        uploaded_file = SimpleUploadedFile(
            "rsvp.csv",
            (
                "Timestamp,Full Name,Student ID,Email,Dietary Restrictions\n"
                "2026-05-01 09:00,Alice Kim,U1234567,alice@example.com,None\n"
            ).encode("utf-8"),
            content_type="text/csv",
        )

        self.client.post("/checkin/import/", {"rsvp_file": uploaded_file})
        self.client.post(
            "/checkin/import/review/",
            {
                "review_action": "confirm",
                "unique_identifier_selection": "Student ID",
                "name_column": "Full Name",
                "major_column": "",
                "email_column": "Email",
                "timestamp_column": "Timestamp",
                "display_columns": ["Full Name", "Email", "Dietary Restrictions"],
                "searchable_columns": ["Full Name", "Email"],
            },
        )

        response = self.client.get("/checkin/registered/")

        self.assertContains(response, "Full Name")
        self.assertContains(response, "Email")
        self.assertContains(response, "Dietary Restrictions")
        self.assertNotContains(response, "Major</th>", html=False)
        self.assertContains(response, 'placeholder="Search by Full Name, Email"', html=False)
        self.assertContains(response, "Export Full Attendance Report (.xlsx)")

    def test_import_review_requires_pending_upload(self):
        response = self.client.get("/checkin/import/review/")

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], "/checkin/import/")


class AttendanceExportViewTests(TestCase):
    def test_export_includes_checked_in_registered_and_all_guests(self):
        RegisteredParticipant.objects.create(
            submission_order=1,
            name="Checked In Student",
            unid="u2000",
            major="Biology",
            checked_in=True,
        )
        RegisteredParticipant.objects.create(
            submission_order=2,
            name="Not Checked In Student",
            unid="u2001",
            major="Physics",
            checked_in=False,
        )
        GuestParticipant.objects.create(
            name="Walk-in Guest",
            unid="g3000",
            major="Visitor",
        )

        response = self.client.get("/checkin/export/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "text/csv")
        content = response.content.decode("utf-8")
        self.assertIn("Registered,Checked In Student,u2000,Biology,Yes,", content)
        self.assertNotIn("Not Checked In Student", content)
        self.assertIn("Guest,Walk-in Guest,g3000,Visitor,Yes,", content)

    def test_rsvp_export_includes_all_registered_participants(self):
        RegisteredParticipant.objects.create(
            submission_order=1,
            name="First Student",
            unid="u2100",
            major="Biology",
            checked_in=True,
        )
        RegisteredParticipant.objects.create(
            submission_order=2,
            name="Second Student",
            unid="u2101",
            major="Physics",
            checked_in=False,
        )

        response = self.client.get("/checkin/export-rsvp/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "text/csv")
        content = response.content.decode("utf-8")
        self.assertIn("Submission Order,Name,UNID,Major,Checked In,Check-in Time", content)
        self.assertIn("1,First Student,u2100,Biology,Yes,", content)
        self.assertIn("2,Second Student,u2101,Physics,No,", content)

    def test_rsvp_xlsx_export_returns_full_attendance_report(self):
        RSVPImportConfiguration.objects.create(
            pk=1,
            imported_columns=["Full Name", "Original Team"],
            display_columns=["Full Name", "Original Team"],
            searchable_columns=["Full Name", "Email"],
            unique_identifier_source="Student ID",
            name_column="Full Name",
            email_column="Email",
        )
        RegisteredParticipant.objects.create(
            submission_order=1,
            name="Alice Kim",
            unid="u2100",
            checked_in=True,
            checkin_time=timezone.now(),
            answers={
                "Full Name": "Alice Kim",
                "Original Team": "Blue Team",
            },
        )
        RegisteredParticipant.objects.create(
            submission_order=2,
            name="Brian Lee",
            unid="u2101",
            checked_in=False,
            answers={
                "Full Name": "Brian Lee",
                "Original Team": "Gold Team",
            },
        )
        GuestParticipant.objects.create(
            name="Walk In Guest",
            unid="g3000",
            major="Visitor",
            checked_in=True,
            checkin_time=timezone.now(),
        )

        response = self.client.get("/checkin/export/xlsx/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(
            workbook.sheetnames,
            [
                "Summary",
                "Registered Participants",
                "Checked-In Only",
                "No-Show",
                "Guest Participants",
                "Final Attendance",
                "Data Analysis",
            ],
        )

        summary_sheet = workbook["Summary"]
        summary_values = {
            summary_sheet[f"A{row}"].value: summary_sheet[f"B{row}"].value
            for row in range(2, summary_sheet.max_row + 1)
        }
        self.assertEqual(summary_values["Total RSVP"], 2)
        self.assertEqual(summary_values["Checked-in RSVP"], 1)
        self.assertEqual(summary_values["No-show RSVP"], 1)
        self.assertEqual(summary_values["Guest Count"], 1)
        self.assertEqual(summary_values["Total Attendance"], 2)
        self.assertEqual(summary_values["Selected imported column names"], "Full Name, Original Team")

        registered_sheet = workbook["Registered Participants"]
        registered_headers = [cell.value for cell in registered_sheet[1]]
        self.assertEqual(
            registered_headers,
            [
                "Unique Identifier",
                "Full Name",
                "Original Team",
                "Check-In Time",
                "Imported At",
                "Check-In Status",
            ],
        )
        self.assertNotIn("Email", registered_headers)
        self.assertNotIn("Notes", registered_headers)
        registered_rows = list(registered_sheet.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(registered_rows), 2)
        self.assertEqual(registered_rows[0][0], "u2100")
        self.assertEqual(registered_rows[0][1], "Alice Kim")
        self.assertEqual(registered_rows[0][2], "Blue Team")
        self.assertEqual(registered_rows[0][-1], "Checked In")
        self.assertEqual(registered_rows[1][-1], "No Show")

        checked_in_sheet = workbook["Checked-In Only"]
        not_checked_in_sheet = workbook["No-Show"]
        self.assertEqual(checked_in_sheet.max_row, 2)
        self.assertEqual(not_checked_in_sheet.max_row, 2)
        self.assertEqual(checked_in_sheet["A2"].value, "u2100")
        self.assertEqual(not_checked_in_sheet["A2"].value, "u2101")

        guest_sheet = workbook["Guest Participants"]
        self.assertEqual(guest_sheet.max_row, 2)
        self.assertEqual(guest_sheet["A2"].value, "Walk In Guest")
        self.assertEqual(guest_sheet["B2"].value, "g3000")
        self.assertEqual(guest_sheet["F2"].value, "Checked In")

        final_attendance_sheet = workbook["Final Attendance"]
        final_headers = [cell.value for cell in final_attendance_sheet[1]]
        self.assertEqual(
            final_headers,
            [
                "Type",
                "Unique Identifier",
                "Full Name",
                "Original Team",
                "Check-In Time",
                "Attendance Status",
            ],
        )
        final_rows = list(final_attendance_sheet.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(final_rows), 2)
        self.assertEqual(final_rows[0][0], "Registered")
        self.assertEqual(final_rows[0][1], "u2100")
        self.assertEqual(final_rows[0][2], "Alice Kim")
        self.assertEqual(final_rows[0][-1], "Present")
        self.assertEqual(final_rows[1][0], "Guest")
        self.assertEqual(final_rows[1][1], "g3000")
        self.assertEqual(final_rows[1][2], "Walk In Guest")
        self.assertEqual(final_rows[1][-1], "Present")

        analysis_rows = list(workbook["Data Analysis"].iter_rows(values_only=True))
        self.assertIn(("Original Team", "RSVP Count", "Checked-in Count", "No-show Count"), analysis_rows)
        self.assertIn(("Blue Team", 1, 1, 0), analysis_rows)
        self.assertIn(("Gold Team", 1, 0, 1), analysis_rows)

    def test_rsvp_xlsx_export_falls_back_to_answers_json_keys_without_display_columns(self):
        RegisteredParticipant.objects.create(
            submission_order=1,
            name="Generated Participant",
            unid="rsvp-00001",
            checked_in=False,
            answers={
                "Will you attend?": "Yes",
                "Dietary Restrictions": "None",
                "Table": "A1",
            },
        )

        response = self.client.get("/checkin/export/xlsx/")

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        headers = [cell.value for cell in workbook["Registered Participants"][1]]
        self.assertIn("Will you attend?", headers)
        self.assertIn("Dietary Restrictions", headers)
        self.assertIn("Table", headers)
        self.assertIn("Unique Identifier", headers)
        self.assertIn("Check-In Status", headers)
        row_values = [cell.value for cell in workbook["Registered Participants"][2]]
        self.assertIn("Yes", row_values)
        self.assertIn("No Show", row_values)
        self.assertIn("A1", row_values)

        analysis_rows = list(workbook["Data Analysis"].iter_rows(values_only=True))
        self.assertIn(("Table", "RSVP Count", "Checked-in Count", "No-show Count"), analysis_rows)
        self.assertIn(("A1", 1, 0, 1), analysis_rows)


class AnalyticsViewTests(TestCase):
    def test_analytics_counts_checked_in_registered_and_guests_by_major(self):
        RegisteredParticipant.objects.create(
            submission_order=1,
            name="Checked In IS Student",
            unid="u3001",
            major="IS",
            checked_in=True,
            checkin_time=timezone.now(),
        )
        RegisteredParticipant.objects.create(
            submission_order=2,
            name="No Show IS Student",
            unid="u3002",
            major="IS",
            checked_in=False,
        )
        RegisteredParticipant.objects.create(
            submission_order=3,
            name="Checked In Finance Student",
            unid="u3003",
            major="Finance",
            checked_in=True,
            checkin_time=timezone.now(),
        )
        GuestParticipant.objects.create(
            name="Walk In IS Guest",
            unid="g3001",
            major="IS",
            checked_in=True,
            checkin_time=timezone.now(),
        )
        GuestParticipant.objects.create(
            name="Walk In Other Guest",
            unid="g3002",
            major="Other",
            checked_in=True,
            checkin_time=timezone.now(),
        )

        response = self.client.get("/checkin/analytics/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(json.loads(response.context["major_labels"]), ["IS", "Finance", "Other"])
        self.assertEqual(json.loads(response.context["major_values"]), [2, 1, 1])
        self.assertEqual(response.context["attendance_group_label"], "Major")
        self.assertContains(response, "Current Attendance by Major")
        self.assertContains(
            response,
            "Checked-in registered participants plus guest check-ins",
        )


class GuestCheckInFormTests(TestCase):
    def test_guest_form_uses_uac_major_list(self):
        RegisteredParticipant.objects.create(
            submission_order=1,
            name="Imported Student",
            unid="u2999999",
            major="Business",
        )

        form = GuestCheckInForm()

        self.assertEqual(
            list(form.fields["major"].widget.choices),
            [("", "Select Major")] + [(major, major) for major in DEFAULT_GUEST_MAJORS],
        )

    def test_guest_form_maps_unlisted_major_to_other(self):
        form = GuestCheckInForm(
            data={
                "name": "Walk In",
                "unid": "u1234567",
                "major": "Business",
            }
        )

        self.assertTrue(form.is_valid())
        self.assertEqual(form.cleaned_data["major"], "Other")


class ParticipantListAndAdminFlowTests(TestCase):
    def test_toggle_checkin_returns_partial_for_htmx_requests(self):
        participant = RegisteredParticipant.objects.create(
            submission_order=1,
            name="HTMX Toggle Student",
            unid="u4000",
            major="Operations",
            checked_in=False,
            checkin_time=None,
        )

        response = self.client.post(
            f"/checkin/participants/{participant.id}/toggle-checkin/",
            HTTP_HX_REQUEST="true",
        )

        participant.refresh_from_db()
        self.assertEqual(response.status_code, 200)
        self.assertTrue(participant.checked_in)
        self.assertContains(response, 'aria-pressed="true"', html=False)
        self.assertNotContains(response, "tap to undo")
        self.assertContains(response, 'hx-swap-oob="true"', html=False)
        self.assertNotIn("Location", response)

    def test_toggle_checkin_updates_status_and_time(self):
        participant = RegisteredParticipant.objects.create(
            submission_order=1,
            name="Toggle Student",
            unid="u4001",
            major="Operations",
            checked_in=False,
            checkin_time=None,
        )

        response = self.client.post(f"/checkin/participants/{participant.id}/toggle-checkin/")

        participant.refresh_from_db()
        self.assertEqual(response.status_code, 302)
        self.assertTrue(participant.checked_in)
        self.assertIsNotNone(participant.checkin_time)

    def test_toggle_checkin_can_uncheck_and_clear_time(self):
        participant = RegisteredParticipant.objects.create(
            submission_order=1,
            name="Toggle Back Student",
            unid="u4003",
            major="Operations",
            checked_in=True,
            checkin_time=timezone.now(),
        )

        response = self.client.post(f"/checkin/participants/{participant.id}/toggle-checkin/")

        participant.refresh_from_db()
        self.assertEqual(response.status_code, 302)
        self.assertFalse(participant.checked_in)
        self.assertIsNone(participant.checkin_time)

    def test_registered_page_does_not_include_undo_helper_or_scroll_restore_hook(self):
        participant = RegisteredParticipant.objects.create(
            submission_order=1,
            name="Stable UI Student",
            unid="u4002",
            major="Operations",
            checked_in=True,
            checkin_time=timezone.now(),
        )

        response = self.client.get("/checkin/registered/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, participant.name)
        self.assertNotContains(response, "tap to undo")
        self.assertNotContains(response, "pendingCheckinScrollY")
        self.assertContains(response, 'focus-scroll:false', html=False)

    def test_delete_participant_removes_row_and_reorders_following_rows(self):
        first = RegisteredParticipant.objects.create(
            submission_order=1,
            name="First Student",
            unid="u5001",
            major="Finance",
        )
        second = RegisteredParticipant.objects.create(
            submission_order=2,
            name="Second Student",
            unid="u5002",
            major="Marketing",
        )
        third = RegisteredParticipant.objects.create(
            submission_order=3,
            name="Third Student",
            unid="u5003",
            major="Operations",
        )

        response = self.client.post(f"/checkin/participants/{second.id}/delete/")

        self.assertEqual(response.status_code, 302)
        self.assertTrue(RegisteredParticipant.objects.filter(id=first.id).exists())
        self.assertFalse(RegisteredParticipant.objects.filter(id=second.id).exists())
        third.refresh_from_db()
        self.assertEqual(third.submission_order, 2)

    def test_delete_all_participants_clears_rsvp_list(self):
        RegisteredParticipant.objects.create(
            submission_order=1,
            name="First Student",
            unid="u6001",
            major="Accounting",
        )
        RegisteredParticipant.objects.create(
            submission_order=2,
            name="Second Student",
            unid="u6002",
            major="Finance",
        )

        response = self.client.post("/checkin/participants/delete-all/")

        self.assertEqual(response.status_code, 302)
        self.assertEqual(RegisteredParticipant.objects.count(), 0)

    def test_import_page_shows_guest_records_and_admin_actions(self):
        GuestParticipant.objects.create(
            name="Walk In Guest",
            unid="u8123000",
            major="Visitor",
            checked_in=True,
            checkin_time=timezone.now(),
        )

        response = self.client.get("/checkin/import/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Guest Records")
        self.assertContains(response, "Walk In Guest")
        self.assertContains(response, "Delete Entire RSVP List")
        self.assertContains(response, "Delete Guest Records")
        self.assertContains(response, "Export Full Attendance Report (.xlsx)")


class DashboardAndCheckInFlowTests(TestCase):
    def test_dashboard_shows_registered_checked_in_and_guest_counts(self):
        RegisteredParticipant.objects.create(
            submission_order=1,
            name="Registered Student One",
            unid="u9001",
            major="Accounting",
            checked_in=True,
            checkin_time=timezone.now(),
        )
        RegisteredParticipant.objects.create(
            submission_order=2,
            name="Registered Student Two",
            unid="u9002",
            major="Finance",
            checked_in=False,
        )
        GuestParticipant.objects.create(
            name="Walk In Guest",
            unid="u9003",
            major="Other",
            checked_in=True,
            checkin_time=timezone.now(),
        )

        response = self.client.get("/checkin/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["total_rsvp"], 2)
        self.assertEqual(response.context["registered_checked_in"], 1)
        self.assertEqual(response.context["guest_count"], 1)
        self.assertEqual(response.context["current_total_attendance"], 2)
        self.assertContains(response, "University of Utah")
        self.assertContains(response, "Current Total Attendance")

    def test_dashboard_handles_unavailable_database_without_500(self):
        with patch("checkin.views.RegisteredParticipant.objects.count", side_effect=OperationalError):
            response = self.client.get("/checkin/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Database setup required")
        self.assertContains(response, "DATABASE_URL")

    def test_guest_checkin_creates_guest_and_redirects_to_dashboard(self):
        response = self.client.post(
            "/checkin/guest/",
            {
                "name": "Guest Visitor",
                "unid": "U7654321",
                "major": "Other",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], "/checkin/")
        self.assertEqual(GuestParticipant.objects.count(), 1)

        guest = GuestParticipant.objects.get()
        self.assertEqual(guest.name, "Guest Visitor")
        self.assertEqual(guest.unid, "u7654321")
        self.assertEqual(guest.major, "Other")
        self.assertTrue(guest.checked_in)
        self.assertIsNotNone(guest.checkin_time)

    def test_guest_checkin_rejects_registered_unid(self):
        RegisteredParticipant.objects.create(
            submission_order=1,
            name="Registered Student",
            unid="u9999999",
            major="Physics",
        )

        response = self.client.post(
            "/checkin/guest/",
            {
                "name": "Duplicate Guest",
                "unid": "u9999999",
                "major": "Physics",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(GuestParticipant.objects.count(), 0)
        self.assertContains(
            response,
            "This UNID is already in the registered RSVP list. Use Registered Check-In instead.",
        )
